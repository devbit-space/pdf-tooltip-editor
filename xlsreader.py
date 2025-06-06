import openpyxl
import os
import re

def read_excel_data(excel_file):
    """Read data from Excel file and return a dictionary of row data"""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active
    
    # Dictionary to store row data
    row_data = {}
    
    for row_idx in range(2, sheet.max_row + 1):
        row = {}
        # Get all non-empty cells in the row
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col)
            if cell.value is not None:
                # Get column header
                header_cell = sheet.cell(row=1, column=col)
                header = str(header_cell.value) if header_cell.value else f"Column_{col}"
                row[header] = str(cell.value)
        
        if row:  # Only add non-empty rows
            row_data[row_idx] = row  # Use row_idx directly to match the Excel row numbers
            
    return row_data

def get_tooltip_content(row_data, tooltip_text):
    """Extract row number and field from tooltip text and get corresponding data"""
    # Extract row number from tooltip text
    if not tooltip_text:
        return None
        
    # Try to find row number in the tooltip text
    row_num = None
    field_name = None
    original_row_text = None
    
    # Clean up the tooltip text
    tooltip_text = tooltip_text.strip().lower()
    
    # Extract the original row text and number
    if 'row' in tooltip_text:
        # Find the complete row reference (e.g., "row 38" or "rows 47-50")
        row_numbers = re.findall(r'row[s]?\s+(\d+(?:-\d+)?)', tooltip_text.lower())
        if row_numbers:
            original_row_text = row_numbers[0]  # Get just the number part
            
        parts = tooltip_text.split('row')
        if len(parts) > 1:
            # Try to extract row number
            for part in parts[1].split():
                try:
                    row_num = int(part.strip("'"))
                    break
                except ValueError:
                    continue
    
    # Common field mappings
    field_mappings = {
        'messstelle': ['messstelle', 'messstelle_mstnr'],
        'gewaessername': ['gewaessername', 'gewässername'],
        'messstellenbezeichnung': ['messstellenbezeichnung'],
        'datum': ['datum'],
        'taxon': ['taxon'],
        'wuchsform': ['wuchsform'],
    }
    
    # Try to find matching field
    for key, variations in field_mappings.items():
        if any(var in tooltip_text for var in variations):
            field_name = key
            break
    
    if row_num and row_num in row_data:
        row = row_data[row_num]
        # If we found a specific field, try to return its value
        if field_name and field_name in row:
            new_content = row[field_name]
        else:
            # If no specific field found or matched, return all row data formatted
            new_content = "\n".join(f"{k}: {v}" for k, v in row.items())
            
        # Format the row information with "row: number"
        if original_row_text:
            if '-' in original_row_text:  # Handle range case like "47-50"
                return f"rows: {original_row_text}\n{new_content}"
            else:
                return f"row: {original_row_text}\n{new_content}"
        
        return new_content
    
    return None

def main():
    # File paths
    excel_file = "Perla_11_1_vorlage-steps.xlsx"
    
    try:
        # Read Excel data
        print("Reading Excel data...")
        row_data = read_excel_data(excel_file)
        
        # Print the data for verification
        print("\nExtracted row data:")
        for row_num, data in row_data.items():
            print(f"\nRow {row_num}:")
            for key, value in data.items():
                print(f"  {key}: {value}")
            
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()